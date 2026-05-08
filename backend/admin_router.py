from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel
from database import get_db
import models
from auth import get_current_user
from ads_store import add_uploaded_ad, remove_ads, set_ad_enabled
import os
import io
import datetime
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

router = APIRouter(prefix="/api/admin", tags=["admin"])

ADMIN_EMAIL = "devanand.s2008@gmail.com"

def check_admin(user: models.User):
    if user.role != "Admin" or user.email != ADMIN_EMAIL:
        raise HTTPException(status_code=403, detail="Not authorized. Admin access required.")


@router.get("/astrologers")
def list_astrologers(token: str, db: Session = Depends(get_db)):
    user = get_current_user(token, db)
    check_admin(user)
    
    astrologers = db.query(models.User).filter(models.User.role == "Astrologer").all()
    return [{"id": a.id, "name": a.name, "email": a.email, "status": a.status, "picture": a.picture} for a in astrologers]

@router.post("/astrologers/{astrologer_id}/approve")
def approve_astrologer(astrologer_id: int, token: str, db: Session = Depends(get_db)):
    user = get_current_user(token, db)
    check_admin(user)
    
    astro = db.query(models.User).filter(models.User.id == astrologer_id).first()
    if not astro:
        raise HTTPException(status_code=404, detail="Astrologer not found")
    
    astro.status = "Approved"
    db.commit()
    return {"message": "Astrologer approved successfully"}

@router.post("/astrologers/{astrologer_id}/reject")
def reject_astrologer(astrologer_id: int, token: str, db: Session = Depends(get_db)):
    user = get_current_user(token, db)
    check_admin(user)
    
    astro = db.query(models.User).filter(models.User.id == astrologer_id).first()
    if not astro:
        raise HTTPException(status_code=404, detail="Astrologer not found")
    
    astro.status = "Rejected"
    db.commit()
    return {"message": "Astrologer rejected successfully"}

# ── Unified User Management (used by admin.html User Mgmt section) ──

@router.get("/users")
def list_all_users(token: str, db: Session = Depends(get_db)):
    """List ALL users for admin management panel."""
    user = get_current_user(token, db)
    check_admin(user)
    users = db.query(models.User).order_by(models.User.created_at.desc()).all()
    return [{"id": u.id, "name": u.name, "email": u.email, "role": u.role,
             "status": u.status, "picture": u.picture or ""} for u in users]

class StatusUpdate(BaseModel):
    token: str
    status: str  # "Approved", "Rejected", "Pending"

@router.post("/users/{user_id}/status")
def update_user_status(user_id: int, req: StatusUpdate, db: Session = Depends(get_db)):
    """Update any user's status — approve/reject astrologers, suspend users."""
    admin = get_current_user(req.token, db)
    check_admin(admin)
    if req.status not in ("Approved", "Rejected", "Pending"):
        raise HTTPException(status_code=400, detail="Invalid status")
    target = db.query(models.User).filter(models.User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    target.status = req.status
    db.commit()
    return {"user_id": user_id, "status": req.status, "name": target.name}


@router.get("/stats")
def get_stats(token: str, db: Session = Depends(get_db)):
    user = get_current_user(token, db)
    check_admin(user)
    
    total_users = db.query(models.User).filter(models.User.role == "User").count()
    total_astrologers = db.query(models.User).filter(models.User.role == "Astrologer").count()
    total_messages = db.query(models.Message).count()
    
    return {
        "total_users": total_users,
        "total_astrologers": total_astrologers,
        "total_messages": total_messages
    }

UPLOADS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads", "ads")
os.makedirs(UPLOADS_DIR, exist_ok=True)

@router.post("/ads/upload")
async def upload_ad(
    token: str = Form(...),
    ad_type: str = Form(...), # "web", "pdf", or "banner"
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    user = get_current_user(token, db)
    check_admin(user)

    try:
        ad = add_uploaded_ad(ad_type, file)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid ad type")

    return {
        "message": "Ad uploaded successfully",
        "path": ad["path"],
        "ad": ad,
    }

@router.delete("/ads/remove")
async def remove_ad(
    token: str,
    ad_type: Optional[str] = None,
    ad_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    user = get_current_user(token, db)
    check_admin(user)

    if not ad_type and not ad_id:
        raise HTTPException(status_code=400, detail="ad_type or ad_id is required")

    removed = remove_ads(ad_type=ad_type, ad_id=ad_id)
    if removed:
        return {"message": "Ad removed", "removed": len(removed)}
    return {"message": "No ad found to remove", "removed": 0}


class AdToggle(BaseModel):
    token: str
    enabled: bool


@router.post("/ads/{ad_id}/toggle")
async def toggle_ad(ad_id: str, req: AdToggle, db: Session = Depends(get_db)):
    user = get_current_user(req.token, db)
    check_admin(user)
    ad = set_ad_enabled(ad_id, req.enabled)
    if not ad:
        raise HTTPException(status_code=404, detail="Ad not found")
    return {"message": "Ad updated", "ad": ad}


# ─────────────────────────────────────────────────────────────────
#  EXCEL EXPORT — all user data as .xlsx
# ─────────────────────────────────────────────────────────────────

@router.get("/export/users")
def export_users_excel(token: str, db: Session = Depends(get_db)):
    """Export all user data as a styled Excel (.xlsx) file."""
    if not _OPENPYXL_OK:
        raise HTTPException(status_code=503, detail="openpyxl not installed. Run: pip install openpyxl")
    admin = get_current_user(token, db)
    check_admin(admin)

    users = db.query(models.User).order_by(models.User.created_at.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jyotish Users"

    # ── Styles ──────────────────────────────────────────────────
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", fgColor="3B1F6B")          # dark purple
    alt_fill      = PatternFill("solid", fgColor="F4F0FF")           # light lavender
    center_align  = Alignment(horizontal="center", vertical="center")
    left_align    = Alignment(horizontal="left",   vertical="center")
    thin          = Side(style="thin", color="CCCCCC")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Status colours
    STATUS_COLORS = {
        "Approved": "1D8C68",
        "Pending":  "C8A84B",
        "Rejected": "E04B4A",
    }

    # ── Headers ──────────────────────────────────────────────────
    COLS = [
        ("ID",          8),
        ("பெயர் (Name)",  22),
        ("Email",         30),
        ("Role",          14),
        ("Status",        12),
        ("பதிவு தேதி (Registered)", 22),
        ("Messages Sent", 16),
        ("Messages Received", 18),
    ]

    for col_idx, (header, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    # ── Data rows ────────────────────────────────────────────────
    for row_idx, u in enumerate(users, start=2):
        sent     = db.query(func.count(models.Message.id)).filter(models.Message.sender_id   == u.id).scalar() or 0
        received = db.query(func.count(models.Message.id)).filter(models.Message.receiver_id == u.id).scalar() or 0

        row_data = [
            u.id,
            u.name or "",
            u.email or "",
            u.role or "",
            u.status or "",
            u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "",
            sent,
            received,
        ]

        fill = alt_fill if row_idx % 2 == 0 else None

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = border
            cell.alignment = center_align if isinstance(value, int) else left_align
            if fill:
                cell.fill = fill
            # Colour-code the Status column (col 5)
            if col_idx == 5 and value in STATUS_COLORS:
                cell.font = Font(bold=True, color=STATUS_COLORS[value])

        ws.row_dimensions[row_idx].height = 18

    # ── Summary sheet ────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14

    summary_data = [
        ("📊 Jyotish 3.0 — User Report",   ""),
        ("Generated At",  __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("",              ""),
        ("Total Users",   db.query(models.User).filter(models.User.role == "User").count()),
        ("Total Astrologers", db.query(models.User).filter(models.User.role == "Astrologer").count()),
        ("Pending Astrologers", db.query(models.User).filter(models.User.role == "Astrologer", models.User.status == "Pending").count()),
        ("Approved Astrologers", db.query(models.User).filter(models.User.role == "Astrologer", models.User.status == "Approved").count()),
        ("Total Admins",  db.query(models.User).filter(models.User.role == "Admin").count()),
        ("Total Messages", db.query(models.Message).count()),
    ]

    for r, (label, val) in enumerate(summary_data, start=1):
        a = ws2.cell(row=r, column=1, value=label)
        b = ws2.cell(row=r, column=2, value=val)
        if r == 1:
            a.font = Font(bold=True, size=13, color="3B1F6B")
        elif r == 2:
            a.font = Font(italic=True, color="888888")
            b.font = Font(italic=True, color="888888")
        else:
            a.font = Font(bold=True)
            b.font = Font(bold=False, color="1D8C68")
        a.alignment = left_align
        b.alignment = left_align
        ws2.row_dimensions[r].height = 20

    # ── Stream response ──────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"jyotish_users_{__import__('datetime').datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
